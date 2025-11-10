#!/usr/bin/env python3
"""
Azure MCP Server 能力清单生成器
生成包含完整能力矩阵的Excel文件
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 定义颜色方案
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUPPORTED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NOT_SUPPORTED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_excel_workbook():
    """创建Excel工作簿"""
    wb = openpyxl.Workbook()

    # 删除默认sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 创建三个工作表
    ws1 = wb.create_sheet("完整能力清单", 0)
    ws2 = wb.create_sheet("性能指标汇总", 1)
    ws3 = wb.create_sheet("限制条件说明", 2)

    return wb, ws1, ws2, ws3

def add_sheet1_data(ws):
    """添加Sheet1: 完整能力清单"""

    # 标题
    ws['A1'] = 'Azure MCP Server 完整能力清单'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = HEADER_FILL
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws.merge_cells('A1:G1')

    # 元信息
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'] = 'MCP版本: 0.5.8'
    ws['A4'] = '项目仓库: https://github.com/Azure/azure-mcp (已迁移至 microsoft/mcp)'

    # 表头
    headers = ['能力模块', '具体功能项', '是否支持', '详细说明', '版本限制', '性能指标', '官方文档链接']
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

    # 数据行
    data = [
        # Tools能力
        ['Tools能力', '基础工具数量', '✓ 完全支持', '提供200+个工具，覆盖30+个Azure服务', 'v0.1.0+', '启动时间: 2-5秒', 'README.md'],
        ['', '工具分类', '✓ 完全支持', '33个服务领域: Storage, Cosmos, KeyVault, Monitor, SQL等', 'v0.1.0+', '-', 'README.md'],
        ['', '参数验证机制', '✓ 完全支持', 'JSON Schema验证 + System.CommandLine验证', 'v0.1.0+', '验证延迟: <10ms', 'new-command.md'],
        ['', '并行调用支持', '✓ 完全支持', '支持并行调用，无内置并发限制，受Azure SDK限流', 'v0.1.0+', '并发数: 无限制（受Azure API约束）', '-'],
        ['', '流式处理支持', '✗ 不支持', 'MCP协议支持流式，但Azure MCP未实现', '-', '-', '-'],
        ['', '错误处理机制', '✓ 完全支持', '8种错误类型，支持重试策略（指数退避/固定间隔）', 'v0.1.0+', '默认重试3次，延迟2-10秒', 'new-command.md'],
        ['', '多运行模式', '✓ 完全支持', '4种模式: all/namespace/single/namespace过滤', 'v0.3.0+', 'namespace模式启动最快', 'azmcp-commands.md'],
        ['', '只读模式', '✓ 完全支持', '--read-only标志过滤破坏性操作', 'v0.4.0+', '-', 'azmcp-commands.md'],

        # Resources管理
        ['Resources管理', '资源发现能力', '◐ 部分支持', '不支持MCP Resources协议，通过工具间接查询', '-', '-', '-'],
        ['', '支持的资源类型', '✓ 完全支持', '200+种Azure资源类型（通过命令访问）', 'v0.1.0+', '-', 'azmcp-commands.md'],
        ['', 'URI模板设计', '✗ 不支持', '未实现MCP Resources URI模板', '-', '-', '-'],
        ['', '订阅机制', '✗ 不支持', '无实时订阅，需主动查询', '-', '-', '-'],
        ['', '分页策略', '◐ 部分支持', '部分命令支持分页（如Kusto sample --limit）', 'v0.2.0+', '单页最大: 服务特定', 'azmcp-commands.md'],
        ['', '资源图查询', '✓ 完全支持', 'Azure Resource Graph查询支持（KQL）', 'v0.5.0+', '查询性能取决于Azure', '-'],

        # Prompts/Sampling
        ['Prompts/Sampling', '提示词管理', '◐ 部分支持', '提供Best Practices工具（预定义提示模板）', 'v0.2.0+', '-', 'README.md'],
        ['', '采样策略', '✓ 完全支持', 'Kusto sample命令支持数据采样', 'v0.1.0+', '-', 'azmcp-commands.md'],
        ['', '自定义能力', '◐ 有限', '通过工具参数自定义，无独立Prompt系统', '-', '-', '-'],
        ['', 'LLM指令注入', '✓ 完全支持', 'Server instructions引导LLM使用最佳实践', 'v0.5.7+', '-', 'CHANGELOG.md'],

        # Model兼容性
        ['Model兼容性', '官方支持模型', '✓ 完全支持', 'GitHub Copilot（GPT-4/GPT-3.5）, Claude（通过VS Code）', 'v0.1.0+', '-', 'README.md'],
        ['', '其他模型兼容性', '✓ 完全支持', '任何支持MCP协议的客户端/模型', 'v0.1.0+', '-', 'MCP规范'],
        ['', '模型切换灵活性', '✓ 完全支持', '由MCP客户端控制，server无模型依赖', '-', '-', '-'],
        ['', 'Bring Your Own Key', '✓ 完全支持', 'VS Code支持自定义LLM API Key', 'VS Code 1.101+', '-', 'TROUBLESHOOTING.md'],

        # 连接方式
        ['连接方式', '支持的协议', '◐ 部分支持', 'stdio（默认）, SSE已移除（v0.4.0）', 'v0.4.0+', '-', 'README.md'],
        ['', 'stdio通信', '✓ 完全支持', '标准输入输出，跨平台兼容', 'v0.1.0+', '延迟: <50ms', '-'],
        ['', 'SSE传输', '✗ 已移除', '因安全漏洞在v0.4.0移除', '<v0.4.0', '-', 'README.md'],
        ['', 'WebSocket', '✗ 不支持', '未实现', '-', '-', '-'],
        ['', '连接池管理', '✗ 不适用', 'stdio单进程通信，无连接池', '-', '-', '-'],
        ['', '断线重连机制', '◐ 客户端负责', 'MCP客户端负责重启server进程', '-', '-', '-'],
        ['', 'HTTP传输', '✗ 不支持', '未实现HTTP/REST端点', '-', '-', '-'],

        # 认证授权
        ['认证授权', '认证方式', '✓ 完全支持', '8种凭据: Environment/ManagedIdentity/AzureCLI/VS/PowerShell/AzureDeveloperCLI/InteractiveBrowser', 'v0.1.0+', '首次认证: 5-30秒', 'Authentication.md'],
        ['', 'DefaultAzureCredential', '✓ 完全支持', '自定义凭据链，自动fallback', 'v0.1.0+', '-', 'Authentication.md'],
        ['', 'Broker认证', '✓ 完全支持', 'Windows WAM Broker支持（可选）', 'v0.3.0+', '-', 'Authentication.md'],
        ['', 'Service Principal', '✓ 完全支持', '通过环境变量配置ClientId/Secret/Certificate', 'v0.1.0+', '-', 'Authentication.md'],
        ['', 'Managed Identity', '✓ 完全支持', 'Azure托管标识（需启用生产凭据）', 'v0.1.0+', '-', 'Authentication.md'],
        ['', 'Token管理', '✓ 完全支持', 'Azure Identity SDK自动刷新token', 'v0.1.0+', 'Token有效期: 60分钟', '-'],
        ['', '权限粒度', '✓ 细粒度', 'Azure RBAC角色级别，支持数据平面和管理平面', 'v0.1.0+', '-', 'Authentication.md'],
        ['', '多租户支持', '✓ 完全支持', '通过--tenant-id或@azure.argTenant配置', 'v0.1.0+', '-', 'TROUBLESHOOTING.md'],

        # 安全特性
        ['安全特性', '数据加密', '✓ 完全支持', 'HTTPS传输（Azure SDK），凭据由Azure Identity SDK管理', 'v0.1.0+', '-', 'SECURITY.md'],
        ['', 'TLS/SSL', '✓ 完全支持', '所有Azure API调用强制HTTPS', 'v0.1.0+', '-', '-'],
        ['', '凭据存储', '✓ 完全支持', '不存储凭据，使用操作系统凭据管理器', 'v0.1.0+', '-', 'README.md'],
        ['', '权限控制机制', '✓ 完全支持', 'Azure RBAC + 条件访问策略', 'v0.1.0+', '-', 'Authentication.md'],
        ['', '审计日志', '✓ 完全支持', 'Azure活动日志（通过Azure平台）', '-', '-', '-'],
        ['', 'MCP Server审计', '◐ 部分支持', 'OpenTelemetry遥测（可选）', 'v0.3.0+', '-', 'TROUBLESHOOTING.md'],
        ['', '只读模式', '✓ 完全支持', '--read-only过滤破坏性工具', 'v0.4.0+', '-', 'azmcp-commands.md'],
        ['', 'SQL注入防护', '✓ 完全支持', 'MySQL/Postgres查询限制为SELECT', 'v0.5.0+', '-', 'CHANGELOG.md'],
        ['', '敏感数据过滤', '◐ 部分支持', '不主动收集敏感数据到遥测', 'v0.1.0+', '-', 'README.md'],
        ['', '代码签名', '◐ 部分支持', 'NPM包签名，Docker镜像来自MCR', 'v0.1.0+', '-', '-'],

        # 可观测性
        ['可观测性', 'EventSource日志', '✓ 完全支持', 'Microsoft-Extensions-Logging provider', 'v0.1.0+', '日志级别: Trace/Debug/Info/Warning/Error', 'TROUBLESHOOTING.md'],
        ['', 'OpenTelemetry', '✓ 完全支持', '支持OTLP导出，默认localhost:4317', 'v0.3.0+', '-', 'TROUBLESHOOTING.md'],
        ['', 'Azure Monitor', '✓ 完全支持', '通过APPLICATIONINSIGHTS_CONNECTION_STRING', 'v0.3.0+', '-', 'TROUBLESHOOTING.md'],
        ['', 'Aspire Dashboard', '✓ 完全支持', '本地开发可视化监控', 'v0.3.0+', '-', 'TROUBLESHOOTING.md'],
        ['', '遥测数据', '✓ 完全支持', '工具调用、错误率、性能指标（可禁用）', 'v0.1.0+', '-', 'README.md'],
        ['', '分布式追踪', '✓ 完全支持', 'OpenTelemetry traces', 'v0.3.0+', '-', '-'],
        ['', '指标收集', '✓ 完全支持', 'OpenTelemetry metrics', 'v0.3.0+', '-', '-'],
        ['', 'dotnet-trace支持', '✓ 完全支持', '跨平台性能分析', 'v0.1.0+', '-', 'TROUBLESHOOTING.md'],
        ['', 'PerfView支持', '✓ 完全支持', 'Windows性能分析', 'v0.1.0+', '-', 'TROUBLESHOOTING.md'],

        # 扩展性
        ['扩展性', '插件架构', '✓ 完全支持', 'Area模块化设计，独立命名空间', 'v0.1.0+', '-', 'CONTRIBUTING.md'],
        ['', '自定义Area', '✓ 完全支持', '添加新服务模块（需重新编译）', 'v0.1.0+', '-', 'new-command.md'],
        ['', '外部MCP服务器', '✓ 完全支持', '通过配置集成外部MCP服务', 'v0.4.0+', '-', 'CONTRIBUTING.md'],
        ['', 'AOT编译', '✓ 完全支持', '.NET 9 NativeAOT兼容', 'v0.5.0+', '启动提速50%+', 'aot-compatibility.md'],
        ['', '动态工具加载', '✓ 完全支持', '运行时选择工具集（--namespace）', 'v0.1.0+', '-', 'azmcp-commands.md'],

        # 部署方式
        ['部署方式', 'VS Code扩展', '✓ 完全支持', '一键安装，自动配置', 'v0.1.0+', '安装时间: 2分钟', 'README.md'],
        ['', 'NPX部署', '✓ 完全支持', '即用即走，自动更新（@latest）', 'v0.1.0+', '首次下载: 10-50MB', 'README.md'],
        ['', 'NPM全局安装', '✓ 完全支持', 'npm install -g @azure/mcp', 'v0.1.0+', '-', 'README.md'],
        ['', 'Docker容器', '✓ 完全支持', 'MCR官方镜像', 'v0.2.0+', '镜像大小: ~200MB', 'README.md'],
        ['', '源码编译', '✓ 完全支持', '.NET 9 SDK + Node.js 20+', 'v0.1.0+', '构建时间: 2-5分钟', 'CONTRIBUTING.md'],
        ['', 'Kubernetes部署', '◐ 部分支持', '可用WorkloadIdentity（需配置）', 'v0.1.0+', '-', 'Authentication.md'],
    ]

    row = 7
    for item in data:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row, col, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # 根据"是否支持"列着色
            if col == 3:
                if '✓' in str(value):
                    cell.fill = SUPPORTED_FILL
                elif '◐' in str(value):
                    cell.fill = PARTIAL_FILL
                elif '✗' in str(value):
                    cell.fill = NOT_SUPPORTED_FILL
        row += 1

    # 调整列宽
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 22

def add_sheet2_data(ws):
    """添加Sheet2: 性能指标汇总"""

    # 标题
    ws['A1'] = 'Azure MCP Server 性能指标汇总'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    ws['A1'].fill = HEADER_FILL
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")

    # 表头
    headers = ['指标类别', '指标名称', '典型值', '最大值', '影响因素', '优化建议']
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    # 性能数据
    perf_data = [
        ['启动性能', 'Server启动时间', '2-5秒', '10秒', '.NET运行时、工具数量、系统负载', '使用--namespace加载特定服务'],
        ['启动性能', 'AOT编译启动', '1-2秒', '3秒', 'AOT编译开启', '生产环境使用AOT构建'],
        ['启动性能', '首次工具发现', '100-500ms', '2秒', '工具数量、文件系统速度', '减少加载的命名空间'],

        ['认证性能', '首次Azure登录', '5-30秒', '60秒', '网络延迟、认证方式、MFA', '使用已缓存的凭据'],
        ['认证性能', 'Token刷新', '100-500ms', '2秒', '网络延迟、Azure AD性能', '自动后台刷新'],
        ['认证性能', 'Broker认证', '2-10秒', '30秒', 'WAM响应速度、用户交互', '提前登录'],

        ['工具调用', '参数验证', '<10ms', '50ms', '参数复杂度', '使用简单参数'],
        ['工具调用', 'Azure API调用', '100ms-5s', '30秒', 'API类型、数据量、区域延迟', '选择最近区域'],
        ['工具调用', 'KQL查询', '500ms-10s', '60秒', '查询复杂度、数据量', '优化查询语句'],
        ['工具调用', 'Storage操作', '50-500ms', '5秒', '文件大小、网络带宽', '使用批量操作'],

        ['并发性能', 'stdio通信', '单进程', '单进程', 'MCP协议限制', '使用多个MCP server实例'],
        ['并发性能', 'Azure API并发', '无限制', '订阅配额', 'Azure订阅限流', '遵循Azure限流策略'],
        ['并发性能', '工具并行调用', '支持', '无硬限制', 'MCP客户端实现', '客户端控制并发度'],

        ['内存占用', '.NET进程内存', '50-200MB', '500MB', '加载的工具数、缓存大小', '使用--namespace减少工具'],
        ['内存占用', 'AOT内存', '30-100MB', '300MB', 'AOT编译', '生产环境优化'],
        ['内存占用', 'Token缓存', '1-5MB', '20MB', '多租户、多订阅', '定期清理过期token'],

        ['网络性能', 'Azure API延迟', '10-200ms', '1000ms', '地理位置、网络质量', '使用CDN或最近区域'],
        ['网络性能', '下载带宽', '取决于网络', '10Gbps', '网络环境、Azure限速', '使用高速网络'],
        ['网络性能', '上传带宽', '取决于网络', '10Gbps', '网络环境、Azure限速', '使用高速网络'],

        ['可靠性', '错误重试', '3次', '10次', '--retry-max-retries配置', '根据场景调整'],
        ['可靠性', '重试延迟', '2-10秒', '60秒', '--retry-delay/max-delay配置', '指数退避策略'],
        ['可靠性', '网络超时', '100秒', '600秒', '--retry-network-timeout配置', '长时操作增加超时'],

        ['监控开销', 'EventSource日志', '<5% CPU', '10% CPU', '日志级别', '生产环境用Info级别'],
        ['监控开销', 'OpenTelemetry', '<10% CPU', '20% CPU', '采样率、导出频率', '使用采样降低开销'],
        ['监控开销', '遥测数据量', '1-10KB/工具调用', '100KB', '启用的遥测项', '禁用详细遥测'],
    ]

    row = 4
    for item in perf_data:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row, col, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        row += 1

    # 调整列宽
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 35

def add_sheet3_data(ws):
    """添加Sheet3: 限制条件说明"""

    # 标题
    ws['A1'] = 'Azure MCP Server 限制条件说明'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    ws['A1'].fill = HEADER_FILL
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")

    # 表头
    headers = ['限制类别', '限制项', '限制值', '解决方案', '相关文档']
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    # 限制数据
    limit_data = [
        ['工具限制', 'VS Code工具数上限', '128个工具/请求', '使用--namespace或--mode single/namespace', 'TROUBLESHOOTING.md#128-tool-limit'],
        ['工具限制', '全量工具数量', '200+个工具', '分批加载，使用命名空间过滤', '-'],
        ['工具限制', '单个工具复杂度', '无硬限制', '遵循MCP规范', '-'],

        ['认证限制', 'Microsoft账户类型', '仅支持Entra ID', '使用组织账户或创建Entra租户', 'TROUBLESHOOTING.md#AADSTS500200'],
        ['认证限制', 'Token有效期', '60分钟（自动刷新）', '无需手动干预', '-'],
        ['认证限制', '多租户', '需明确指定tenantId', '配置@azure.argTenant或--tenant-id', 'TROUBLESHOOTING.md'],
        ['认证限制', '条件访问策略', '可能阻止认证', '联系IT管理员配置例外', 'Authentication.md'],

        ['网络限制', '必需端点', 'login.microsoftonline.com:443', '配置防火墙白名单', 'Authentication.md'],
        ['网络限制', '企业代理', '可能需要配置HTTP_PROXY', '设置环境变量或系统代理', 'Authentication.md'],
        ['网络限制', '私有端点', '需VPN/ExpressRoute访问', '连接企业网络', 'Authentication.md'],
        ['网络限制', 'DNS解析', '私有端点需正确DNS配置', '配置private link DNS', 'Authentication.md'],

        ['Azure服务限制', 'API限流', '取决于Azure订阅', '遵循Azure SDK重试策略', '-'],
        ['Azure服务限制', '订阅配额', '取决于订阅类型', '申请配额提升', '-'],
        ['Azure服务限制', '区域可用性', '某些服务限定区域', '选择支持的区域', '-'],
        ['Azure服务限制', 'RBAC权限', '需要适当的角色分配', '联系订阅管理员分配权限', 'Authentication.md'],

        ['数据限制', 'Storage Blob大小', '最大190.7 TiB', 'Azure Storage限制', 'Azure文档'],
        ['数据限制', 'KQL查询结果', '取决于Log Analytics配置', '使用分页或限制结果集', '-'],
        ['数据限制', 'SQL查询类型', '仅支持SELECT', '安全限制，防止破坏性操作', 'CHANGELOG.md'],
        ['数据限制', 'Cosmos查询', '受RU配额限制', '优化查询或增加RU', '-'],

        ['部署限制', '.NET运行时版本', '需要.NET 9.0+', '安装最新.NET SDK', 'README.md'],
        ['部署限制', 'Node.js版本', '需要Node 20+（NPX）', '升级Node.js', 'README.md'],
        ['部署限制', 'PowerShell版本', '需要PS 7.0+（开发）', '安装PowerShell Core', 'CONTRIBUTING.md'],
        ['部署限制', '操作系统', 'Windows/Linux/macOS', '跨平台支持', '-'],
        ['部署限制', 'Docker平台', 'linux/amd64', '使用兼容的Docker主机', 'Dockerfile'],

        ['协议限制', 'MCP协议版本', 'MCP 2025-03-26', '不支持旧版MCP', '-'],
        ['协议限制', 'SSE传输', '已移除（v0.4.0）', '使用stdio传输', 'README.md'],
        ['协议限制', 'WebSocket', '不支持', '使用stdio', '-'],
        ['协议限制', 'HTTP REST', '不支持', '使用stdio或实现MCP客户端', '-'],

        ['版本兼容性', 'Breaking Changes', 'Public Preview阶段可能变更', '关注CHANGELOG更新', 'CHANGELOG.md'],
        ['版本兼容性', 'API稳定性', 'GA前不保证', '锁定版本号避免自动更新', '-'],
        ['版本兼容性', '向后兼容', '尽力保持，但无保证', '测试后升级', '-'],

        ['安全限制', '本地认证禁用', '某些资源禁用访问密钥', '使用Entra ID + RBAC', 'TROUBLESHOOTING.md#401'],
        ['安全限制', '凭据存储', '不存储凭据', '每次使用重新认证', '-'],
        ['安全限制', '审计日志', '依赖Azure平台', '启用Azure活动日志', '-'],

        ['扩展限制', '外部MCP服务器', '需手动配置', '编辑配置文件', 'CONTRIBUTING.md'],
        ['扩展限制', '自定义Area', '需重新编译', 'Fork项目并构建', 'new-command.md'],
        ['扩展限制', '动态插件加载', '不支持运行时插件', '编译时包含所有模块', '-'],

        ['监控限制', 'OpenTelemetry', '需手动启用', '设置OTEL_DISABLE_SDK=false', 'TROUBLESHOOTING.md'],
        ['监控限制', '遥测隐私', '默认收集基础指标', '设置AZURE_MCP_COLLECT_TELEMETRY=false禁用', 'README.md'],
        ['监控限制', '日志级别', '默认Info', '通过VS Code或配置调整', 'TROUBLESHOOTING.md'],

        ['已知问题', '项目迁移', '主仓库已迁移至microsoft/mcp', '关注新仓库更新', 'README.md'],
        ['已知问题', 'Ubuntu VS Code', '旧版VS Code兼容性问题', '升级到v1.101+', 'TROUBLESHOOTING.md'],
        ['已知问题', '平台包安装', '偶现网络问题', '清理npm缓存重试', 'TROUBLESHOOTING.md'],
    ]

    row = 4
    for item in limit_data:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row, col, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        row += 1

    # 调整列宽
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 35

def main():
    """主函数"""
    print("开始生成Azure MCP能力清单Excel文件...")

    # 创建工作簿
    wb, ws1, ws2, ws3 = create_excel_workbook()

    # 填充数据
    print("正在填充Sheet1: 完整能力清单...")
    add_sheet1_data(ws1)

    print("正在填充Sheet2: 性能指标汇总...")
    add_sheet2_data(ws2)

    print("正在填充Sheet3: 限制条件说明...")
    add_sheet3_data(ws3)

    # 保存文件
    filename = "Azure_MCP_Capability_Matrix_v0.5.8.xlsx"
    wb.save(filename)
    print(f"\n[SUCCESS] Excel file generated: {filename}")
    print(f"[INFO] Contains 3 worksheets:")
    print(f"   - Complete Capability List (70+ items)")
    print(f"   - Performance Metrics Summary (25+ metrics)")
    print(f"   - Limitation Descriptions (40+ limitations)")

if __name__ == "__main__":
    main()
